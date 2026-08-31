"""Tests for the openpyxl-based Excel helpers."""

from __future__ import annotations

import io

import openpyxl
import pytest

from apps.shared.excel import build_workbook, workbook_to_bytes


@pytest.mark.unit
class TestBuildWorkbook:
    def test_single_sheet_with_headers_and_rows(self):
        workbook = build_workbook(
            sheets=[
                (
                    "Students",
                    ["NISN", "Name"],
                    [("1234567890", "Budi"), ("9876543210", "Ani")],
                ),
            ]
        )

        assert isinstance(workbook, openpyxl.Workbook)
        assert workbook.sheetnames == ["Students"]

        rows = list(workbook["Students"].iter_rows(values_only=True))
        assert rows[0] == ("NISN", "Name")
        assert rows[1] == ("1234567890", "Budi")
        assert rows[2] == ("9876543210", "Ani")

    def test_multiple_sheets_in_order(self):
        workbook = build_workbook(
            sheets=[
                ("First", ["h"], [("a",)]),
                ("Second", ["h"], [("b",)]),
                ("Third", ["h"], [("c",)]),
            ]
        )
        assert workbook.sheetnames == ["First", "Second", "Third"]

    def test_empty_rows_produces_header_only_sheet(self):
        workbook = build_workbook(sheets=[("Empty", ["h1", "h2"], [])])
        rows = list(workbook["Empty"].iter_rows(values_only=True))
        assert rows == [("h1", "h2")]

    def test_default_sheet_is_removed(self):
        # openpyxl's Workbook() constructor creates an active sheet named "Sheet"
        # by default. build_workbook should remove it so we never produce
        # workbooks with a stray default sheet.
        workbook = build_workbook(sheets=[("Only", ["h"], [])])
        assert "Sheet" not in workbook.sheetnames


@pytest.mark.unit
class TestWorkbookToBytes:
    def test_returns_bytes(self):
        workbook = build_workbook(sheets=[("S", ["h"], [])])
        result = workbook_to_bytes(workbook)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_round_trip(self):
        workbook = build_workbook(sheets=[("S", ["a", "b"], [(1, 2), (3, 4)])])
        buffer = workbook_to_bytes(workbook)

        reloaded = openpyxl.load_workbook(filename=io.BytesIO(buffer))
        rows = list(reloaded["S"].iter_rows(values_only=True))
        assert rows[0] == ("a", "b")
        assert rows[1] == (1, 2)
        assert rows[2] == (3, 4)
